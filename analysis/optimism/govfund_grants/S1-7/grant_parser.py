import re
import json
from pathlib import Path
from typing import Dict, Tuple, List, Optional

import pandas as pd
import numpy as np
import openpyxl

# -------------------------------------------------------------------------
# Configuration
# -------------------------------------------------------------------------

# Sheets to skip entirely
SKIP_SHEETS = {"status key", "statuskey"}

# Sheets that have an extra header row to skip (sheet_name -> rows_to_skip)
HEADER_ROW_OVERRIDES: Dict[str, int] = {
    "Grants Season 5": 1,
    "Grants Season 6": 1,
    "Grants Season 7": 1,
}

# Expected output schema
REQUIRED_COLS = [
    "Project Name",
    "Grants Season or Mission",
    "Status",
    "Initial delivery date",
    "OP Delivered",
    "OP Total Amount",
    "Intent",
    "Link",
]
NUMERIC_COLS = ["OP Delivered", "OP Total Amount"]

# Column name variants → standardized names
COLUMN_MAP: Dict[str, str] = {
    'Project Name': 'Project Name',
    'Status': 'Status',
    'Initial Delivery Date': 'Initial delivery date',
    'Delivery Date': 'Initial delivery date',
    '(OP) Delivered': 'OP Delivered',
    'Delivered (OP)': 'OP Delivered',
    'Total Amount (OP)': 'OP Total Amount',
    'Amount (OP)': 'OP Total Amount',
    'Proposal Link': 'Link',
    'Intent ': 'Intent',
    'Intent': 'Intent',
}

# Regex / substrings for summary rows
SUMMARY_RE = re.compile(
    r'^\s*(total\b|season\b|#|%|amount\b|percent\b|proposal|proposals|%\s+of|number\s+of)',
    flags=re.IGNORECASE,
)
SUMMARY_SUBSTRINGS = [
    "total op",
    "total proposals",
    "proposals that received",
    "proposals passed",
    "season total",
    "of proposals",
    "season %",
    "% of op",
    "% of season total",
    "cycle total",
    "token house governance pool",
    "op allocated",
    "season total op approved",
    "# of proposals"
]

# Global variable to store summary data
season_summaries = {}

# Add test links for debugging
TEST_LINKS = [
    "https://app.charmverse.io/op-grants/page-7215315141818317",
    "https://app.charmverse.io/op-grants/page-11908812258760593",
    "https://app.charmverse.io/op-grants/proposals?id=bcaedc5b-fbc5-46af-a9e4-d292bbf1cde2",
    "https://app.charmverse.io/op-grants/page-11674564580610602"
]

# -------------------------------------------------------------------------
# Helpers
# -------------------------------------------------------------------------

def is_summary_row(project_name: str) -> bool:
    """Check if a row is a summary row based on its project name."""
    if pd.isna(project_name):
        return False
    project_name = str(project_name).lower()
    
    # First check if it matches the regex pattern
    if SUMMARY_RE.search(project_name):
        return True
        
    # Then check for specific substrings
    return any(s in project_name for s in SUMMARY_SUBSTRINGS)


def detect_header_row(xlsx: Path, sheet: str, lookahead: int = 5) -> Optional[int]:
    """
    Scan the first `lookahead` rows to find the row that contains 'Project Name'.
    Returns the zero-based header row index, or None if not found.
    """
    peek = pd.read_excel(
        xlsx, sheet_name=sheet, header=None, nrows=lookahead, engine='openpyxl'
    )
    for idx, row in peek.iterrows():
        if row.astype(str).str.contains('Project Name', case=False, na=False).any():
            return idx
    return None


def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Standardize column names across different sheets."""
    
    # Define column mappings
    column_mappings = {
        'Project Name': 'Project Name',
        'Status': 'Status',
        'Initial Delivery Date': 'Initial delivery date',
        'Delivery Date': 'Initial delivery date',
        '(OP) Delivered': 'OP Delivered',
        'Delivered (OP)': 'OP Delivered',
        'Total Amount (OP)': 'OP Total Amount',
        'Amount (OP)': 'OP Total Amount',
        'Proposal Link': 'Proposal Link',
        'Intent ': 'Intent',
        'Intent': 'Intent'
    }
    
    # Create a copy of the dataframe
    df = df.copy()
    
    # Rename columns based on mappings
    renamed_cols = {}
    for old_col in df.columns:
        if old_col in column_mappings:
            renamed_cols[old_col] = column_mappings[old_col]
    
    # Apply the renaming
    df = df.rename(columns=renamed_cols)
    
    return df


# -------------------------------------------------------------------------
# Main ingestion function
# -------------------------------------------------------------------------

def clean_workbook(xlsx_path: str) -> pd.DataFrame:
    # Load the workbook directly with openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    sheets = [
        s for s in wb.sheetnames
        if s.lower() not in {"status key", "statuskey"}
    ]

    frames = []
    summary_data = {}
    
    for sheet in sheets:
        # Read first few rows to determine the correct header
        df_peek = pd.read_excel(
            xlsx_path,
            sheet_name=sheet,
            header=None,
            nrows=5
        )
        
        # Find the row that contains "Project Name" as a header
        header_row = None
        for idx, row in df_peek.iterrows():
            if row.astype(str).str.contains('Project Name').any():
                header_row = idx
                break
                
        if header_row is None:
            print(f"Warning: Could not find header row in {sheet}")
            continue
            
        # Read the sheet with the correct header
        df = pd.read_excel(
            xlsx_path,
            sheet_name=sheet,
            header=header_row
        )
        
        # Handle both Proposal Link column and Project Name hyperlinks
        if 'Project Name' in df.columns:
            ws = wb[sheet]
            
            # Find the Project Name column index
            project_name_col = None
            for cell in ws[header_row + 1]:  # +1 because Excel is 1-based
                if cell.value == 'Project Name':
                    project_name_col = cell.column - 1  # Convert to 0-based index
                    break
            
            if project_name_col is not None:
                # Create a mapping of row indices to hyperlinks
                hyperlinks = {}
                for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 2, max_row=ws.max_row)):
                    cell = row[project_name_col]
                    if hasattr(cell, 'hyperlink') and cell.hyperlink:
                        hyperlinks[row_idx] = cell.hyperlink.target
                
                # Add hyperlinks to Link column where available
                df['Link'] = df.index.map(lambda x: hyperlinks.get(x, pd.NA))
        
        # Handle Proposal Link column
        if 'Proposal Link' in df.columns:
            if sheet == 'Grants Season 6':
                # For Season 6, use Proposal Link only for rows that don't have a Project Name hyperlink
                df.loc[df['Link'].isna(), 'Link'] = df.loc[df['Link'].isna(), 'Proposal Link']
            else:
                # For other sheets, only use Proposal Link to fill in missing links
                df.loc[df['Link'].isna(), 'Link'] = df.loc[df['Link'].isna(), 'Proposal Link']
        
        # Drop duplicate columns before standardization
        df = df.loc[:, ~df.columns.duplicated()].copy()
        df = standardize_columns(df)
        df = df.loc[:, ~df.columns.duplicated()]

        # First, capture summary rows if we have the required columns
        if "Project Name" in df.columns and ("OP Total Amount" in df.columns or "OP Delivered" in df.columns):
            summary_rows = df[
                df["Project Name"].fillna("").apply(is_summary_row)
            ].copy()
            
            if not summary_rows.empty:
                total_approved = None
                total_allocated = None
                
                for _, row in summary_rows.iterrows():
                    project_name = str(row["Project Name"]).lower()
                    if "total" in project_name and "op" in project_name and "approved" in project_name:
                        total_approved = pd.to_numeric(row.get("OP Total Amount", pd.NA), errors='coerce')
                    elif "amount" in project_name and "op" in project_name and "allocated" in project_name:
                        total_allocated = pd.to_numeric(row.get("OP Delivered", pd.NA), errors='coerce')
                
                if total_approved is not None or total_allocated is not None:
                    summary_data[sheet] = {
                        "Total OP Approved": total_approved,
                        "Amount of OP Allocated": total_allocated
                    }

            # Now process regular rows
            df = df[
                ~df["Project Name"].fillna("").apply(is_summary_row)
            ].copy()

        # Ensure required columns exist
        for col in REQUIRED_COLS:
            if col not in df.columns:
                df[col] = pd.NA

        # Add sheet label
        df["Grants Season or Mission"] = sheet

        # Capture extra columns as JSON "Other metadata"
        meta_cols = [
            c for c in df.columns 
            if c not in REQUIRED_COLS + ["Grants Season or Mission"]
        ]
        df["Other metadata"] = df[meta_cols].apply(
            lambda row: json.dumps(
                {k: v for k, v in row.items() if pd.notna(v) and str(v) != "nan"}
            ),
            axis=1,
        )

        # Select only the required columns plus metadata, ensuring no duplicates
        final_cols = REQUIRED_COLS + ["Grants Season or Mission", "Other metadata"]
        df = df.loc[:, final_cols].copy()
        
        # Ensure all columns are present and in the correct order
        for col in final_cols:
            if col not in df.columns:
                df[col] = pd.NA
        df = df[final_cols]
            
        # Convert numeric columns
        for col in NUMERIC_COLS:
            if col in df.columns:
                # Convert to string first to handle both numeric and string inputs
                df[col] = df[col].astype(str)
                # Remove commas and convert to numeric
                df[col] = pd.to_numeric(df[col].str.replace(',', ''), errors='coerce')
        
        # Set OP Total Amount to 0 for Not-passed projects in Missions Season 4
        if sheet == 'Missions Season 4' and 'OP Total Amount' in df.columns and 'Status' in df.columns:
            df.loc[df['Status'] == 'Not-passed', 'OP Total Amount'] = 0
        
        frames.append(df)

    # Store summary data in a global variable or return it separately
    global season_summaries
    season_summaries = summary_data

    wb.close()

    result = pd.concat(frames, ignore_index=True)
    return result.loc[:, ~result.columns.duplicated()]